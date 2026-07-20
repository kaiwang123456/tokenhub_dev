"""
多轮对话自动标注脚本

功能：
    对 wildchat_multiturn_2000.xlsx 中的多轮对话进行三分类标注（simple/complex/coding）
    基于 classify_prompt.py 的 sp_gpt prompt（多轮对话专用版本）
    使用 Qwen3-4B 作为分类器，输出 JSON（intent + confidence + reason）

特性：
    - 断点续传：已有标签的行自动跳过，中断后再跑不重复标注
    - JSON 解析容错：提取 JSON 块、清理非法字符、重试
    - 定期自动保存：每 BATCH_SAVE 条保存一次
    - 失败重试：单条 API 最多重试 MAX_RETRIES 次
    - 日志文件：完整记录标注过程

用法：
    python auto_label_multiturn.py

配置：
    修改下方 CONFIG 区域可调整 API、模型、路径等
"""

import json
import time
import re
import urllib.request
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ======================== CONFIG ========================
API_URL = "https://wishub-x6.ctyun.cn/v1/chat/completions"
API_KEY = "0ef2d0dbdf334d48b81d528ffd127d3c"
MODEL = "Qwen3-4B"

INPUT_XLSX = r"C:\Users\27942\Desktop\wildchat_multiturn_2000.xlsx"
OUTPUT_XLSX = r"C:\Users\27942\Desktop\wildchat_multiturn_2000_labeled.xlsx"
OUTPUT_JSON = r"C:\Users\27942\Desktop\wildchat_multiturn_2000_labeled.json"
LOG_FILE = r"C:\Users\27942\Desktop\multiturn_label_log.txt"

BATCH_SAVE = 50           # 每标注N条自动保存
REQUEST_INTERVAL = 0.3    # API请求间隔(秒)
MAX_RETRIES = 3           # 单条请求最大重试次数
CONVERSATION_MAX_LEN = 2000  # 发给API的对话文本最大长度
# ========================================================


# ======================== PROMPT ========================
# 基于 classify_prompt.py 的 sp_gpt，为多轮对话设计
# 修复了原版 intent/label 不一致问题，统一使用 intent

SP_GPT = """# Role

You are a task classifier for an AI Gateway.

Your responsibility is to classify the entire conversation into exactly ONE of the following categories:

- simple
- complex
- coding

The classification result will be used for model routing.

Return ONLY valid JSON.

---

# Classification Rules

## 1. coding

Select "coding" if the user's primary objective is software development or source code processing.

This includes but is not limited to:

- writing code
- modifying code
- debugging
- explaining code
- reviewing code
- refactoring
- SQL
- Shell
- Python
- Java
- C++
- Go
- Rust
- JavaScript
- HTML/CSS
- Docker
- Kubernetes
- YAML
- JSON Schema
- API design
- Git
- Regular Expressions
- unit tests
- software architecture for implementation

If software engineering is the main topic, ALWAYS classify as "coding", regardless of reasoning complexity.

---

## 2. complex

Select "complex" if solving the request requires substantial reasoning, planning, analysis, synthesis, or integrating multiple constraints.

Typical characteristics:

- multi-step reasoning
- long-context understanding
- architecture or system design
- project planning
- research
- report generation
- business analysis
- legal analysis
- financial analysis
- comparing multiple solutions
- designing workflows
- generating long structured documents
- decision making with multiple factors

The answer cannot be produced by a straightforward response.

---

## 3. simple

Select "simple" if the request can be completed directly without significant reasoning or planning.

Typical examples:

- factual questions
- definitions
- translation
- grammar correction
- rewriting
- short summarization
- basic calculations
- simple recommendations
- FAQs
- short explanations

Usually requires only a direct response.

---

# Priority

If multiple categories appear applicable, use the following priority:

coding > complex > simple

---

# Conversation Scope

Classify based on the ENTIRE conversation rather than only the last user message.

If earlier turns establish that the conversation is about software development, classify as coding even if the latest message is brief.

---

# Output

Return ONLY JSON, no other text.

{
    "intent": "simple|complex|coding",
    "confidence": 0.00-1.00,
    "reason": "one short sentence"
}

---

# Conversation to classify:

{conversation}"""


# ======================== API CALL ========================

def call_api(conversation, retries=MAX_RETRIES):
    """调用API对对话进行分类，返回原始输出"""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    # 截断对话
    conv_text = conversation[:CONVERSATION_MAX_LEN]
    prompt = SP_GPT.replace("{conversation}", conv_text)

    body = {
        "model": MODEL,
        "messages": [
            {"role": "user", "content": prompt},
        ],
        "max_tokens": 150,
        "enable_thinking": False,
    }
    data = json.dumps(body).encode("utf-8")

    for attempt in range(retries):
        try:
            req = urllib.request.Request(API_URL, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json.loads(resp.read().decode("utf-8"))
                return result["choices"][0]["message"]["content"].strip()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
            else:
                return f"ERROR: {e}"


# ======================== JSON PARSING ========================

def parse_json_response(raw):
    """解析模型输出，容错处理"""
    if raw.startswith("ERROR"):
        return None, raw

    # 尝试直接解析
    try:
        data = json.loads(raw)
        return _extract_label(data), raw
    except:
        pass

    # 尝试提取 JSON 块
    json_match = re.search(r'\{[^{}]*\}', raw, re.DOTALL)
    if json_match:
        try:
            data = json.loads(json_match.group())
            return _extract_label(data), raw
        except:
            pass

    # 尝试提取带嵌套的 JSON
    json_match = re.search(r'\{.*?\}', raw, re.DOTALL)
    if json_match:
        try:
            data = json.loads(json_match.group())
            return _extract_label(data), raw
        except:
            pass

    # 兜底：关键词匹配
    lower = raw.lower()
    if "coding" in lower:
        return "coding", raw
    if "complex" in lower:
        return "complex", raw
    if "simple" in lower:
        return "simple", raw

    return None, raw


def _extract_label(data):
    """从JSON dict提取标签，兼容 intent/label 两种字段名"""
    label = data.get("intent") or data.get("label") or data.get("category") or ""
    label = str(label).strip().lower()
    if label in ("simple", "complex", "coding"):
        return label
    # 兜底
    for key, val in data.items():
        val_str = str(val).strip().lower()
        if val_str in ("simple", "complex", "coding"):
            return val_str
    return None


# ======================== SAVE ========================

def save_results(wb, results_json):
    """保存Excel和JSON"""
    wb.save(OUTPUT_XLSX)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results_json, f, ensure_ascii=False, indent=2)


def log(msg):
    """同时输出到控制台和日志文件"""
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(msg + "\n")


# ======================== MAIN ========================

def main():
    # 初始化日志
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write(f"多轮对话标注日志 - {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"模型: {MODEL}\n")
        f.write(f"输入: {INPUT_XLSX}\n\n")

    log("=" * 60)
    log("多轮对话自动标注脚本")
    log("=" * 60)
    log(f"输入: {INPUT_XLSX}")
    log(f"输出Excel: {OUTPUT_XLSX}")
    log(f"输出JSON: {OUTPUT_JSON}")
    log(f"模型: {MODEL}")
    log(f"每 {BATCH_SAVE} 条自动保存")
    log("")

    # 读取Excel
    log("读取Excel...")
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active
    total_rows = ws.max_row - 1
    log(f"共 {total_rows} 条对话待标注")
    log("")

    # 统计
    labeled_count = 0
    skipped_count = 0
    label_counts = Counter()
    error_count = 0
    unknown_count = 0
    results_json = []
    start_time = time.time()

    for row_idx in range(2, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=1).value
        if no is None:
            continue
        # 跳过非数据行（如AIGC水印行，No列为UUID字符串）
        try:
            int(no)
        except (ValueError, TypeError):
            continue

        lang = ws.cell(row=row_idx, column=2).value or ""
        turn = ws.cell(row=row_idx, column=3).value or ""
        conversation = ws.cell(row=row_idx, column=4).value or ""
        existing_label = ws.cell(row=row_idx, column=5).value

        # 断点续传：已有标签跳过
        if existing_label and str(existing_label).strip().lower() in ("simple", "complex", "coding"):
            label_counts[str(existing_label).strip().lower()] += 1
            skipped_count += 1
            results_json.append({
                "no": int(no),
                "language": lang,
                "turn": turn,
                "conversation": conversation[:200],
                "label": str(existing_label).strip().lower(),
                "method": "skip",
            })
            continue

        # 调用API
        raw_output = call_api(conversation)
        label, raw_clean = parse_json_response(raw_output)

        # 写入Excel
        ws.cell(row=row_idx, column=5).value = label if label else "UNKNOWN"

        labeled_count += 1
        if label is None:
            unknown_count += 1
            label_counts["unknown"] += 1
        else:
            label_counts[label] += 1
        if raw_output.startswith("ERROR"):
            error_count += 1

        results_json.append({
            "no": int(no),
            "language": lang,
            "turn": turn,
            "conversation": conversation[:200],
            "label": label if label else "unknown",
            "raw_output": raw_output[:300],
            "method": "auto",
        })

        # 进度
        if labeled_count % 25 == 0:
            elapsed = time.time() - start_time
            speed = labeled_count / elapsed if elapsed > 0 else 0
            eta = (total_rows - labeled_count - skipped_count) / speed if speed > 0 else 0
            log(
                f"  [{labeled_count + skipped_count}/{total_rows}] "
                f"simple={label_counts['simple']} complex={label_counts['complex']} "
                f"coding={label_counts['coding']} unknown={unknown_count} "
                f"err={error_count} | {speed:.1f}条/s ETA={eta/60:.1f}min"
            )

        # 定期保存
        if labeled_count % BATCH_SAVE == 0:
            save_results(wb, results_json)
            log(f"  >> 自动保存 ({labeled_count + skipped_count}/{total_rows})")

        time.sleep(REQUEST_INTERVAL)

    # 最终保存
    save_results(wb, results_json)
    elapsed = time.time() - start_time

    log("")
    log("=" * 60)
    log("标注完成")
    log("=" * 60)
    log(f"  新标注: {labeled_count} 条")
    log(f"  跳过(已有标签): {skipped_count} 条")
    log(f"  耗时: {elapsed/60:.1f} 分钟")
    log(f"  API错误: {error_count} 条")
    log(f"  UNKNOWN: {unknown_count} 条")
    log(f"")
    log(f"  标签分布:")
    for k in ["simple", "complex", "coding", "unknown"]:
        v = label_counts.get(k, 0)
        total = labeled_count + skipped_count
        if v > 0:
            log(f"    {k}: {v} ({v/total*100:.1f}%)")

    # 按语言统计
    log(f"")
    log(f"  按语言:")
    for lang_name in ["Chinese", "English"]:
        lang_items = [r for r in results_json if r["language"] == lang_name]
        lc = Counter(r["label"] for r in lang_items)
        log(f"    {lang_name}({len(lang_items)}): simple={lc['simple']} complex={lc['complex']} coding={lc['coding']} unknown={lc.get('unknown',0)}")

    log(f"")
    log(f"  Excel: {OUTPUT_XLSX}")
    log(f"  JSON:  {OUTPUT_JSON}")
    log(f"  日志:  {LOG_FILE}")

    # UNKNOWN样本提示
    unknowns = [r for r in results_json if r["label"] == "unknown"]
    if unknowns:
        log(f"")
        log(f"  UNKNOWN样本 ({len(unknowns)}条):")
        for u in unknowns[:10]:
            log(f"    no={u['no']} raw={u.get('raw_output','')[:80]}")


if __name__ == "__main__":
    main()
